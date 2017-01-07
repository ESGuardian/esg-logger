#! /usr/bin/python
# -*- coding: utf8 -*-
import sys
from datetime import date, timedelta, datetime
from pytz import timezone
from elasticsearch import Elasticsearch 
import codecs


def getindexes (es,prefix,dates) :
    indexes = []
    for strdate in dates:
        testindex = prefix + strdate
        if es.indices.exists(index=testindex):
            indexes.append(testindex)
    return indexes
    
period=1
if len(sys.argv) > 1:
    period=int(sys.argv[1])

remote_access = True
app_install = True
srv_install = True
antimalware = True
group = True
accounts = True
ora_accounts = True
ora_access = True
ips = True

dict = {4727:u"Создана глобальная группа с включенной безопасностью.",\
4728:u"Член добавлен к глобальной группе с включенной безопасностью.",\
4729:u"Член удален из глобальной группы с включенной безопасностью.",\
4730:u"Удалена глобальная группа с включенной безопасностью.",\
4754:u"Создана универсальная группа с проверкой безопасности.",\
4756:u"Член добавлен к универсальной группе с проверкой безопасности.",\
4757:u"Член удален из универсальной группы с включенной безопасностью.",\
4758:u"Удалена универсальная группа с проверкой безопасности.",\
4764:u"Изменен тип группы.",\
4731:u"Создана локальная группа с включенной безопасностью.",\
4732:u"Член добавлен к локальной группе с включенной безопасностью.",\
4733:u"Член удален из локальной группы с включенной безопасностью.",\
4734:u"Удалена локальная группа с включенной безопасностью.",\
4720:u"Была создана учетная запись пользователя.",\
4722:u"Учетная запись пользователя была включена.",\
4723:u"Предпринята попытка изменить пароль учетной записи.",\
4724:u"Предпринята попытка выполнить сброс пароля учетной записи.",\
4725:u"Учетная запись пользователя была отключена.",\
4726:u"Учетная запись пользователя была удалена.",\
4740:u"Учетная запись пользователя заблокирована.",\
4767:u"Учетная запись пользователя была разблокирована.",\
4781:u"Было изменено имя учетной записи.",\
5377:u"Диспетчер учетных данных учетные данные были восстановлены из резервной копии.",\
4741:u"Была создана учетная запись компьютера.",\
4743:u"Учетная запись компьютера была удалена."}

today = datetime.utcnow().date()
enddate=today.strftime('%Y.%m.%d')
startdate=(today - timedelta(days=period)).strftime('%Y.%m.%d')
dates = []
for i in xrange(0,period+1):
    dates.append((today - timedelta(days=i)).strftime('%Y.%m.%d'))

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
wb = Workbook()
first = True
es = Elasticsearch()

if remote_access :
    if first :
        ws = wb.active
        ws.title = u'Удаленка'
        first = False
    else : 
        ws = wb.create_sheet(u'Удаленка') 
    row_cursor = 1
    tabheader = u"Удаленный доступ с " + startdate + u" по " + enddate
    c = ws.cell(row=row_cursor, column=1, value=tabheader)
    c.style = 'Title'
    c.alignment = Alignment(horizontal='left',vertical='center')
    ws.merge_cells(start_row=row_cursor,start_column=1,end_row=row_cursor+2,end_column=7)
    row_cursor += 3
    colheader = [u'Время',u'Имя',u'Группа',u'IP',u'Страна',u'Город',u'Назначенный IP']
    i = 0
    for head in colheader:
        i += 1
        c = ws.cell(row=row_cursor, column=i, value=head)
        c.style = 'Headline 2'
    row_cursor += 1
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 20
    
    myquery =   {"query":\
                        {\
                            "constant_score":{ "filter":{"term":{"event-code":"Remconn address assigned"}} }\
                        },\
                        "sort":{"@timestamp":{"order":"asc"}},\
                        "size":"10000"\
                    }


    try:
        res = es.search(index=getindexes(es,"cisco-asa-",dates),body=myquery)
        for hit in res['hits']['hits']:
            timestr = hit["_source"]["@timestamp"].replace("T"," ").split(".")[0]
            timeUTC = datetime.strptime(timestr, "%Y-%m-%d %H:%M:%S")
            timeUTC = timeUTC.replace(tzinfo=timezone('UTC'))
            timeLocal = timeUTC.astimezone(timezone('Europe/Moscow'))
            str = timeLocal.strftime("%Y-%m-%d %H:%M:%S")
            c = ws.cell(row=row_cursor, column=1, value=str)
            c.alignment = Alignment(horizontal='left',vertical='center')
            str = hit["_source"]["user"]
            c = ws.cell(row=row_cursor, column=2, value=str)
            c.alignment = Alignment(horizontal='left',vertical='center')
            str = hit["_source"]["group"]
            c = ws.cell(row=row_cursor, column=3, value=str)
            c.alignment = Alignment(horizontal='left',vertical='center')
            str = hit["_source"]["src_ip"]
            c = ws.cell(row=row_cursor, column=4, value=str)
            c.alignment = Alignment(horizontal='left',vertical='center')
            str = hit["_source"]["geoip"]["country_code2"]
            c = ws.cell(row=row_cursor, column=5, value=str)
            c.alignment = Alignment(horizontal='left',vertical='center')
            try:
                str = hit["_source"]["geoip"]["city_name"]
            except:
                str = u"-"
            c = ws.cell(row=row_cursor, column=6, value=str)
            c.alignment = Alignment(horizontal='left',vertical='center')
            str = hit["_source"]["assigned_ip"]
            c = ws.cell(row=row_cursor, column=7, value=str)
            c.alignment = Alignment(horizontal='left',vertical='center')
            row_cursor += 1
    except Exception as e: 
        print "ERROR:" + unicode(e)
        pass
    
    row_cursor += 1
    tabheader = u"Неудачный доступ с " + startdate + u" по " + enddate
    c = ws.cell(row=row_cursor, column=1, value=tabheader)
    c.style = 'Title'
    c.alignment = Alignment(horizontal='left',vertical='center')
    ws.merge_cells(start_row=row_cursor,start_column=1,end_row=row_cursor+2,end_column=7)
    row_cursor += 3
    colheader = [u'Время',u'Имя',u'Сервер доступа',u'IP',u'Страна',u'Город',u'Причина']
    i = 0
    for head in colheader:
        i += 1
        c = ws.cell(row=row_cursor, column=i, value=head)
        c.style = 'Headline 2'
    row_cursor += 1
    myquery =   {"query":\
                        {\
                            "constant_score":{ "filter":{"term":{"event-code":"AAA user authentication Rejected"}} }\
                        },\
                        "sort":{"@timestamp":{"order":"asc"}},\
                        "size":"10000"\
                    }

    
    try:
        res = es.search(index=getindexes(es,"cisco-asa-",dates),body=myquery)
        for hit in res['hits']['hits']:
            timestr = hit["_source"]["@timestamp"].replace("T"," ").split(".")[0]
            timeUTC = datetime.strptime(timestr, "%Y-%m-%d %H:%M:%S")
            timeUTC = timeUTC.replace(tzinfo=timezone('UTC'))
            timeLocal = timeUTC.astimezone(timezone('Europe/Moscow'))
            str = timeLocal.strftime("%Y-%m-%d %H:%M:%S")
            c = ws.cell(row=row_cursor, column=1, value=str)
            c.alignment = Alignment(horizontal='left',vertical='center')
            try:
                str = hit["_source"]["user"]
            except:
                str = u"-"
            c = ws.cell(row=row_cursor, column=2, value=str)
            c.alignment = Alignment(horizontal='left',vertical='center')
            str = hit["_source"]["auth_server"]
            c = ws.cell(row=row_cursor, column=3, value=str)
            c.alignment = Alignment(horizontal='left',vertical='center')
            str = hit["_source"]["src_ip"]
            c = ws.cell(row=row_cursor, column=4, value=str)
            c.alignment = Alignment(horizontal='left',vertical='center')
            str = hit["_source"]["geoip"]["country_code2"]
            c = ws.cell(row=row_cursor, column=5, value=str)
            c.alignment = Alignment(horizontal='left',vertical='center')
            try:
                str = hit["_source"]["geoip"]["city_name"]
            except:
                str += u"-"
            c = ws.cell(row=row_cursor, column=6, value=str)
            c.alignment = Alignment(horizontal='left',vertical='center')
            str = hit["_source"]["reason"]
            c = ws.cell(row=row_cursor, column=7, value=str)
            c.alignment = Alignment(horizontal='left',vertical='center')
            row_cursor += 1
    except Exception as e: 
        print "ERROR: " + unicode(e)
        pass                
    c = ws.cell(row=row_cursor, column=1, value=u'Проверено: Событий, требующих внимания, нет.')
    c.alignment = Alignment(horizontal='left',vertical='center', wrap_text=True)
    c.font = Font(bold=True)
    ws.merge_cells(start_row=row_cursor,start_column=1,end_row=row_cursor+2,end_column=7)

if group :
    if first :
        ws = wb.active
        ws.title = u'Группы'
        first = False
    else : 
        ws = wb.create_sheet(u'Группы') 
    row_cursor = 1
    tabheader = u"Изменение групп безопасности с " + startdate + u" по " + enddate
    c = ws.cell(row=row_cursor, column=1, value=tabheader)
    c.style = 'Title'
    c.alignment = Alignment(horizontal='left',vertical='center')
    ws.merge_cells(start_row=row_cursor,start_column=1,end_row=row_cursor+2,end_column=7)
    row_cursor += 3
    colheader = [u'Время',u'Действие',u'Источник',u'Оператор',u'Группа',u'Пользователь',u'Подробности']
    i = 0
    for head in colheader:
        i += 1
        c = ws.cell(row=row_cursor, column=i, value=head)
        c.style = 'Headline 2'
    row_cursor += 1
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 30
    ws.column_dimensions['F'].width = 45
    ws.column_dimensions['G'].width = 20
    myquery =   {"query":\
                    {\
                        "constant_score":{ "filter":{"bool":{"must":{"term":{"log_name":"Security"}},"should":{"terms":{ "event_id":[4727,4728,4729,4730,4754,4756,4757,4758,4764,4731,4732,4733,4734] }}}} }\
                    },\
                    "sort":{"@timestamp":{"order":"asc"}},\
                    "size":"10000"\
                }
    
    try:
        res = es.search(index=getindexes(es,"winlogbeat-",dates),body=myquery)
        for hit in res['hits']['hits']:
            timestr = hit["_source"]["@timestamp"].replace("T"," ").split(".")[0]
            timeUTC = datetime.strptime(timestr, "%Y-%m-%d %H:%M:%S")
            timeUTC = timeUTC.replace(tzinfo=timezone('UTC'))
            timeLocal = timeUTC.astimezone(timezone('Europe/Moscow'))
            str = timeLocal.strftime("%Y-%m-%d %H:%M:%S")
            c = ws.cell(row=row_cursor, column=1, value=str)
            c.alignment = Alignment(horizontal='left',vertical='center')
            str = dict[hit["_source"]["event_id"]]
            c = ws.cell(row=row_cursor, column=2, value=str)
            c.alignment = Alignment(horizontal='left',vertical='center', wrap_text=True)
            try:
                str = hit["_source"]["beat"]["hostname"]
            except:
                str = u"-"
            c = ws.cell(row=row_cursor, column=3, value=str)
            c.alignment = Alignment(horizontal='left',vertical='center')
            try:
                str = hit["_source"]["event_data"]["SubjectUserName"]
            except:
                str = u"-"
            c = ws.cell(row=row_cursor, column=4, value=str)
            c.alignment = Alignment(horizontal='left',vertical='center')
            try:
                str = hit["_source"]["event_data"]["TargetUserName"]
            except:
                str = u"-"
            c = ws.cell(row=row_cursor, column=5, value=str)
            c.alignment = Alignment(horizontal='left',vertical='center', wrap_text=True)
            try:
                str = hit["_source"]["event_data"]["MemberName"]
            except:
                str = u'-'
            if str == '-' :
                try:
                    str = hit["_source"]["event_data"]["MemberSid"]
                except:
                    str = u"-"
            c = ws.cell(row=row_cursor, column=6, value=str)
            c.alignment = Alignment(horizontal='left',vertical='center') 
            urlstring = 'http://elastic/app/kibana#/doc/winlogbeat-*/' + hit["_index"] + u'/winlogbeat?id=' + hit["_id"]  
            c = ws.cell(row=row_cursor, column=7, value=u'Подробнее ...')            
            c.style = 'Hyperlink'
            c.alignment = Alignment(horizontal='left',vertical='center')
            c.hyperlink = (urlstring)
            row_cursor += 1
    except Exception as e: 
        print "ERROR: " +  unicode(e)
        pass
    c = ws.cell(row=row_cursor, column=1, value=u'Проверено: Событий, требующих внимания, нет.')
    c.alignment = Alignment(horizontal='left',vertical='center', wrap_text=True)
    c.font = Font(bold=True)
    ws.merge_cells(start_row=row_cursor,start_column=1,end_row=row_cursor+2,end_column=7)

if accounts :
    if first :
        ws = wb.active
        ws.title = u'Юзеры'
        first = False
    else : 
        ws = wb.create_sheet(u'Юзеры') 
    row_cursor = 1
    tabheader = u"\nИзменение учетных записей с " + startdate + u" по " + enddate
    c = ws.cell(row=row_cursor, column=1, value=tabheader)
    c.style = 'Title'
    c.alignment = Alignment(horizontal='left',vertical='center')
    ws.merge_cells(start_row=row_cursor,start_column=1,end_row=row_cursor+2,end_column=6)
    row_cursor += 3
    colheader = [u'Время',u'Действие',u'Источник',u'Оператор',u'Субъект',u'Подробности']
    i = 0
    for head in colheader:
        i += 1
        c = ws.cell(row=row_cursor, column=i, value=head)
        c.style = 'Headline 2'
    row_cursor += 1
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 30
    ws.column_dimensions['F'].width = 20

    myquery =   {"query":\
                    {\
                        "constant_score":{ "filter":{"bool":{"must":{"term":{"log_name":"Security"}},"should":{"terms":{ "event_id":[4720,4722,4723,4724,4725,4726,4740,4767,4781,5377,4741,4743] }}}} }\
                    },\
                    "sort":{"@timestamp":{"order":"asc"}},\
                    "size":"10000"\
                }
    
          
    try:
        res = es.search(index=getindexes(es,"winlogbeat-",dates),body=myquery)
        for hit in res['hits']['hits']:
            timestr = hit["_source"]["@timestamp"].replace("T"," ").split(".")[0]
            timeUTC = datetime.strptime(timestr, "%Y-%m-%d %H:%M:%S")
            timeUTC = timeUTC.replace(tzinfo=timezone('UTC'))
            timeLocal = timeUTC.astimezone(timezone('Europe/Moscow'))
            str = timeLocal.strftime("%Y-%m-%d %H:%M:%S")
            c = ws.cell(row=row_cursor, column=1, value=str)
            c.alignment = Alignment(horizontal='left',vertical='center')
            str = dict[hit["_source"]["event_id"]]
            c = ws.cell(row=row_cursor, column=2, value=str)
            c.alignment = Alignment(horizontal='left',vertical='center', wrap_text=True)
            
            try:
                str = hit["_source"]["beat"]["hostname"]
            except:
                str = u"-"
            c = ws.cell(row=row_cursor, column=3, value=str)
            c.alignment = Alignment(horizontal='left',vertical='center')
            try:
                str = hit["_source"]["event_data"]["SubjectUserName"]
            except:
                str = u"-"
            c = ws.cell(row=row_cursor, column=4, value=str)
            c.alignment = Alignment(horizontal='left',vertical='center')
            try:
                str = hit["_source"]["event_data"]["TargetUserName"]
            except:
                str = u"-"
            c = ws.cell(row=row_cursor, column=5, value=str)
            c.alignment = Alignment(horizontal='left',vertical='center')
            urlstring = 'http://elastic/app/kibana#/doc/winlogbeat-*/' + hit["_index"] + u'/winlogbeat?id=' + hit["_id"]  
            c = ws.cell(row=row_cursor, column=6, value=u'Подробнее ...')            
            c.style = 'Hyperlink'
            c.alignment = Alignment(horizontal='left',vertical='center')
            c.hyperlink = (urlstring)
            row_cursor += 1
    except Exception as e: 
        print "ERROR: " +  unicode(e)
        pass
    c = ws.cell(row=row_cursor, column=1, value=u'Проверено: Событий, требующих внимания, нет.')
    c.alignment = Alignment(horizontal='left',vertical='center', wrap_text=True)
    c.font = Font(bold=True)
    ws.merge_cells(start_row=row_cursor,start_column=1,end_row=row_cursor+2,end_column=6)

if ora_accounts :
    if first :
        ws = wb.active
        ws.title = u'ORACLE'
        first = False
    else : 
        ws = wb.create_sheet(u'ORACLE') 
    row_cursor = 1
    tabheader = u"\nОперации с учетками ORACLE с " + startdate + u" по " + enddate
    c = ws.cell(row=row_cursor, column=1, value=tabheader)
    c.style = 'Title'
    c.alignment = Alignment(horizontal='left',vertical='center')
    ws.merge_cells(start_row=row_cursor,start_column=1,end_row=row_cursor+2,end_column=7)
    row_cursor += 3
    colheader = [u'Время',u'Действие',u'Оператор',u'Хост',u'Объект',u'Субъект',u'Подробности']
    i = 0
    for head in colheader:
        i += 1
        c = ws.cell(row=row_cursor, column=i, value=head)
        c.style = 'Headline 2'
    row_cursor += 1
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 40
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 20
    myquery =   {"query":\
                    {\
                        "constant_score":{\
                            "filter":{"terms":{"action":["51","114","115","7","49","43"]}}\
                        }\
                    },\
                    "sort":{"@timestamp":{"order":"asc"}},\
                    "size":"10000"\
                }
    try:
        res = es.search(index=getindexes(es,"oracle-",dates),body=myquery)
        for hit in res['hits']['hits']:
            timestr = hit["_source"]["timestamp"].replace("T"," ").split(".")[0]
            timeUTC = datetime.strptime(timestr, "%Y-%m-%d %H:%M:%S")
            timeUTC = timeUTC.replace(tzinfo=timezone('UTC'))
            timeLocal = timeUTC.astimezone(timezone('Europe/Moscow'))
            str = timeLocal.strftime("%Y-%m-%d %H:%M:%S")
            c = ws.cell(row=row_cursor, column=1, value=str)
            c.alignment = Alignment(horizontal='left',vertical='center')
            try:
                str = hit["_source"]["action_name"]
            except:
                str = u"-"
            c = ws.cell(row=row_cursor, column=2, value=str)
            c.alignment = Alignment(horizontal='left',vertical='center')
            try:
                str = hit["_source"]["os_username"]
            except:
                str = u"-"
            c = ws.cell(row=row_cursor, column=3, value=str)
            c.alignment = Alignment(horizontal='left',vertical='center')
            try:
                str = hit["_source"]["userhost"]
            except:
                str = u"-"
            c = ws.cell(row=row_cursor, column=4, value=str)
            c.alignment = Alignment(horizontal='left',vertical='center')
            try:
                str = hit["_source"]["obj_name"]
            except:
                str = u"-"    
            c = ws.cell(row=row_cursor, column=5, value=str)
            c.alignment = Alignment(horizontal='left',vertical='center')
            try:
                str = hit["_source"]["grantee"]
            except:
                str = u"-"
            c = ws.cell(row=row_cursor, column=6, value=str)
            c.alignment = Alignment(horizontal='left',vertical='center')
            urlstring = 'http://elastic/app/kibana#/doc/oracle-*/' + hit["_index"] + u'/abs?id=' + hit["_id"]  
            c = ws.cell(row=row_cursor, column=7, value=u'Подробнее ...')            
            c.style = 'Hyperlink'
            c.alignment = Alignment(horizontal='left',vertical='center')
            c.hyperlink = (urlstring)
            row_cursor += 1
    except Exception as e: 
        print "ERROR: " +  unicode(e)
        pass
    c = ws.cell(row=row_cursor, column=1, value=u'Проверено: Событий, требующих внимания, нет.')
    c.alignment = Alignment(horizontal='left',vertical='center', wrap_text=True)
    c.font = Font(bold=True)
    ws.merge_cells(start_row=row_cursor,start_column=1,end_row=row_cursor+2,end_column=7)

if ora_access :
    if first :
        ws = wb.active
        ws.title = u'Доступ к Oracle'
        first = False
    else : 
        ws = wb.create_sheet(u'Доступ к Oracle') 
    row_cursor = 1
    tabheader = u"Подключения к БД Oracle с " + startdate + u" по " + enddate
    c = ws.cell(row=row_cursor, column=1, value=tabheader)
    c.style = 'Title'
    c.alignment = Alignment(horizontal='left',vertical='center')
    ws.merge_cells(start_row=row_cursor,start_column=1,end_row=row_cursor+2,end_column=4)
    row_cursor += 3
    colheader = [u'Пользователь',u'Хост',u'Учетка ORACLE',u'Количество']
    i = 0
    for head in colheader:
        i += 1
        c = ws.cell(row=row_cursor, column=i, value=head)
        c.style = 'Headline 2'
    row_cursor += 1
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 35

    myquery =   {"query":{ "constant_score":{ "filter":{"term":{"action_name":"LOGON"}} } },\
        "sort":[{"@timestamp":{"order":"asc"}}],\
        "size":"10000",\
        "aggs": {\
            "by_osuser": {"terms": {"field": "os_username", "size":200}, "aggs": {"by_userhost": {"terms": {"field": "userhost"}, "aggs":{"by_username": {"terms": {"field": "username"}}}}}}\
        }\
    }
    try:
        res = es.search(index=getindexes(es,"oracle-",dates),body=myquery)        
        for os_user in res['aggregations']['by_osuser']['buckets']:
            os_user_name = os_user['key']
            for user_host in os_user['by_userhost']['buckets']:
                user_host_name = user_host['key']
                for ora_user in user_host['by_username']['buckets']:
                    ora_user_name = ora_user['key']
                    count = ora_user['doc_count']
                    c = ws.cell(row=row_cursor, column=1, value=os_user_name)
                    c.alignment = Alignment(horizontal='left',vertical='center')
                    c = ws.cell(row=row_cursor, column=2, value=user_host_name)
                    c.alignment = Alignment(horizontal='left',vertical='center')
                    c = ws.cell(row=row_cursor, column=3, value=ora_user_name)
                    c.alignment = Alignment(horizontal='left',vertical='center')
                    c = ws.cell(row=row_cursor, column=4, value=count)
                    c.alignment = Alignment(vertical='center')
                    row_cursor += 1
            row_cursor += 1
    except Exception as e: 
        print "ERROR: " +  unicode(e)
        pass
    row_cursor += 2
    
    tabheader = u"Подключения к АБС (неудачные) с " + startdate + u" по " + enddate
    c = ws.cell(row=row_cursor, column=1, value=tabheader)
    c.style = 'Title'
    c.alignment = Alignment(horizontal='left',vertical='center')
    ws.merge_cells(start_row=row_cursor,start_column=1,end_row=row_cursor+2,end_column=4)
    row_cursor += 3
    colheader = [u'Пользователь',u'Хост',u'Учетка ORACLE',u'Количество']
    i = 0
    for head in colheader:
        i += 1
        c = ws.cell(row=row_cursor, column=i, value=head)
        c.style = 'Headline 2'
    row_cursor += 1
    myquery =   {"query":{ "constant_score":{ "filter":{"term":{"action_name":"LOGON FAILURE"}} } },\
        "sort":[{"@timestamp":{"order":"asc"}}],\
        "size":"10000",\
        "aggs": {\
            "by_osuser": {"terms": {"field": "os_username", "size":200}, "aggs": {"by_userhost": {"terms": {"field": "userhost"}, "aggs":{"by_username": {"terms": {"field": "username"}}}}}}\
        }\
    }

    try:
        res = es.search(index=getindexes(es,"oracle-",dates),body=myquery)        
        for os_user in res['aggregations']['by_osuser']['buckets']:
            os_user_name = os_user['key']
            for user_host in os_user['by_userhost']['buckets']:
                user_host_name = user_host['key']
                for ora_user in user_host['by_username']['buckets']:
                    ora_user_name = ora_user['key']
                    count = ora_user['doc_count']
                    c = ws.cell(row=row_cursor, column=1, value=os_user_name)
                    c.alignment = Alignment(horizontal='left',vertical='center')
                    c = ws.cell(row=row_cursor, column=2, value=user_host_name)
                    c.alignment = Alignment(horizontal='left',vertical='center')
                    c = ws.cell(row=row_cursor, column=3, value=ora_user_name)
                    c.alignment = Alignment(horizontal='left',vertical='center')
                    c = ws.cell(row=row_cursor, column=4, value=count)
                    c.alignment = Alignment(vertical='center')
                    row_cursor += 1
            row_cursor += 1
    except Exception as e: 
        print "ERROR: " +  unicode(e)
        pass
    c = ws.cell(row=row_cursor, column=1, value=u'Проверено: Событий, требующих внимания, нет.')
    c.alignment = Alignment(horizontal='left',vertical='center', wrap_text=True)
    c.font = Font(bold=True)
    ws.merge_cells(start_row=row_cursor,start_column=1,end_row=row_cursor+2,end_column=4)
    
if app_install :
    if first :
        ws = wb.active
        ws.title = u'Приложения'
        first = False
    else : 
        ws = wb.create_sheet(u'Приложения') 
    row_cursor = 1
    tabheader = u"Установка приложений с " + startdate + u" по " + enddate
    c = ws.cell(row=row_cursor, column=1, value=tabheader)
    c.style = 'Title'
    c.alignment = Alignment(horizontal='left',vertical='center')
    ws.merge_cells(start_row=row_cursor,start_column=1,end_row=row_cursor+2,end_column=4)
    row_cursor += 3
    colheader = [u'Время',u'Источник',u'Оператор',u'Сообщение']
    i = 0
    for head in colheader:
        i += 1
        c = ws.cell(row=row_cursor, column=i, value=head)
        c.style = 'Headline 2'
    row_cursor += 1
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 100
    
    myquery =   {"query":\
                    {\
                        "constant_score":{ "filter":{"bool":{"must":{"term":{"source_name":"MsiInstaller"}},"should":{"terms":{ "event_id":[1036,1037,1033,1034,11707,1022,1019,1007] }}}} }\
                    },\
                    "sort":{"@timestamp":{"order":"asc"}},\
                    "size":"10000"\
                }

    
    try:
        res = es.search(index=getindexes(es,"winlogbeat-",dates),body=myquery)
        for hit in res['hits']['hits']:
            #alignment=Alignment(horizontal='left', vertical='midle', text_rotation=0, wrap_text=False, shrink_to_fit=True,indent=0)
            timestr = hit["_source"]["@timestamp"].replace("T"," ").split(".")[0]
            timeUTC = datetime.strptime(timestr, "%Y-%m-%d %H:%M:%S")
            timeUTC = timeUTC.replace(tzinfo=timezone('UTC'))
            timeLocal = timeUTC.astimezone(timezone('Europe/Moscow'))
            str = timeLocal.strftime("%Y-%m-%d %H:%M:%S")
            c = ws.cell(row=row_cursor, column=1, value=str)
            c.alignment = Alignment(horizontal='left',vertical='center')
            # str = timeLocal.strftime("%Y-%m-%d %H:%M:%S")
            try:
                str = hit["_source"]["beat"]["hostname"]
            except:
                str = u"-"
            c = ws.cell(row=row_cursor, column=2, value=str)
            c.alignment = Alignment(horizontal='left',vertical='center')
            try:
                str = hit["_source"]["user"]["name"]
            except:
                str = u"-"
            c = ws.cell(row=row_cursor, column=3, value=str)
            c.alignment = Alignment(horizontal='left',vertical='center')
            try:
                str = hit["_source"]["message"].strip().replace(","," ").replace(";",":") 
            except:
                str = u"-"
            c = ws.cell(row=row_cursor, column=4, value=str)
            c.alignment = Alignment(horizontal='left',vertical='center', wrap_text=True)
            row_cursor += 1
    except Exception as e: 
        print "ERROR: " +  unicode(e)
        pass
    c = ws.cell(row=row_cursor, column=1, value=u'Проверено: Событий, требующих внимания, нет.')
    c.alignment = Alignment(horizontal='left',vertical='center', wrap_text=True)
    c.font = Font(bold=True)
    ws.merge_cells(start_row=row_cursor,start_column=1,end_row=row_cursor+4,end_column=4)

if srv_install :
    if first :
        ws = wb.active
        ws.title = u'Сервисы'
        first = False
    else : 
        ws = wb.create_sheet(u'Сервисы') 
    row_cursor = 1
    tabheader = u"\nУстановка сервисов с " + startdate + u" по " + enddate
    c = ws.cell(row=row_cursor, column=1, value=tabheader)
    c.style = 'Title'
    c.alignment = Alignment(horizontal='left',vertical='center')
    ws.merge_cells(start_row=row_cursor,start_column=1,end_row=row_cursor+2,end_column=7)
    row_cursor += 3
    colheader = [u"Время",u"Действие",u"Источник",u"Учетка сервиса",u"Название",u"Запускаемый файл",u"Подробности"]
    i = 0
    for head in colheader:
        i += 1
        c = ws.cell(row=row_cursor, column=i, value=head)
        c.style = 'Headline 2'
    row_cursor += 1
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 40
    ws.column_dimensions['F'].width = 60
    ws.column_dimensions['G'].width = 20
    myquery =   {"query":\
                    {\
                        "constant_score":{ "filter":{"bool":{"must":{"term":{"log_name":"Security"}},"should":{"terms":{ "event_id":[4697] }}}} }\
                    },\
                    "sort":{"@timestamp":{"order":"asc"}},\
                    "size":"10000"\
                }

    try:
        res = es.search(index=getindexes(es,"winlogbeat-",dates),body=myquery)
        for hit in res['hits']['hits']:
            timestr = hit["_source"]["@timestamp"].replace("T"," ").split(".")[0]
            timeUTC = datetime.strptime(timestr, "%Y-%m-%d %H:%M:%S")
            timeUTC = timeUTC.replace(tzinfo=timezone('UTC'))
            timeLocal = timeUTC.astimezone(timezone('Europe/Moscow'))
            str = timeLocal.strftime("%Y-%m-%d %H:%M:%S")
            c = ws.cell(row=row_cursor, column=1, value=str)
            c.alignment = Alignment(horizontal='left',vertical='center')
            event_id = hit["_source"]["event_id"]
            if event_id == 4697:
                str = u"В системе установлена служба."
            else:
                str = u"-"
            c = ws.cell(row=row_cursor, column=2, value=str)
            c.alignment = Alignment(horizontal='left',vertical='center')
            try:
                str = hit["_source"]["beat"]["hostname"]
            except:
                str = u"-"
            c = ws.cell(row=row_cursor, column=3, value=str)
            c.alignment = Alignment(horizontal='left',vertical='center')
            try:
                str = hit["_source"]["event_data"]["ServiceAccount"]
            except:
                str = u"-"
            c = ws.cell(row=row_cursor, column=4, value=str)
            c.alignment = Alignment(horizontal='left',vertical='center')
            try:
                str = hit["_source"]["event_data"]["ServiceName"]
            except:
                str = u"-"
            c = ws.cell(row=row_cursor, column=5, value=str)
            c.alignment = Alignment(horizontal='left',vertical='center')
            try:
                str = hit["_source"]["event_data"]["ServiceFileName"]
            except:
                str = u"-"
            c = ws.cell(row=row_cursor, column=6, value=str)
            c.alignment = Alignment(horizontal='left',vertical='center')
            
            urlstring = 'http://elastic/app/kibana#/doc/winlogbeat-*/' + hit["_index"] + u'/winlogbeat?id=' + hit["_id"]  
            c = ws.cell(row=row_cursor, column=7, value=u'Подробнее ...')            
            c.style = 'Hyperlink'
            c.alignment = Alignment(horizontal='left',vertical='center')
            c.hyperlink = (urlstring)
            row_cursor += 1
    except Exception as e: 
        print "ERROR: " +  unicode(e)
        pass
    c = ws.cell(row=row_cursor, column=1, value=u'Проверено: Событий, требующих внимания, нет.')
    c.alignment = Alignment(horizontal='left',vertical='center', wrap_text=True)
    c.font = Font(bold=True)
    ws.merge_cells(start_row=row_cursor,start_column=1,end_row=row_cursor+2,end_column=6)

if antimalware :
    if first :
        ws = wb.active
        ws.title = u'Антивирус'
        first = False
    else : 
        ws = wb.create_sheet(u'Антивирус') 
    row_cursor = 1
    tabheader = u"\nСобытия Endpoint Protection с " + startdate + u" по " + enddate
    c = ws.cell(row=row_cursor, column=1, value=tabheader)
    c.style = 'Title'
    c.alignment = Alignment(horizontal='left',vertical='center')
    ws.merge_cells(start_row=row_cursor,start_column=1,end_row=row_cursor+2,end_column=5)
    row_cursor += 3
    colheader = [u'Время',u'Код сбытия',u'Источник',u'Сообщение',u'Подробности']
    i = 0
    for head in colheader:
        i += 1
        c = ws.cell(row=row_cursor, column=i, value=head)
        c.style = 'Headline 2'
    row_cursor += 1
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 80
    ws.column_dimensions['E'].width = 20
    myquery =   {"query":\
                    {\
                        "constant_score":{ "filter":{"bool":{"must":{"term":{"source_name":"Microsoft Antimalware"} }, "must_not":{"terms":{"level":["Сведения","Information"]} } } } }\
                    },\
                    "sort":{"@timestamp":{"order":"asc"}},\
                    "size":"10000"\
                }


    try:
        res = es.search(index=getindexes(es,"winlogbeat-",dates),body=myquery)
        for hit in res['hits']['hits']:
            timestr = hit["_source"]["@timestamp"].replace("T"," ").split(".")[0]
            timeUTC = datetime.strptime(timestr, "%Y-%m-%d %H:%M:%S")
            timeUTC = timeUTC.replace(tzinfo=timezone('UTC'))
            timeLocal = timeUTC.astimezone(timezone('Europe/Moscow'))
            str = timeLocal.strftime("%Y-%m-%d %H:%M:%S")
            c = ws.cell(row=row_cursor, column=1, value=str)
            c.alignment = Alignment(horizontal='left',vertical='center')
            try:
                id = hit["_source"]["event_id"]
            except:
                id = u"-"
            c = ws.cell(row=row_cursor, column=2, value=id)
            c.alignment = Alignment(horizontal='left',vertical='center')
            if id == 1116 :
                for i in xrange(1,6):
                    ws.cell(row=row_cursor, column=i).fill = PatternFill("solid", fgColor="FFFF00")
            try:
                str = hit["_source"]["beat"]["hostname"]
            except:
                str = u"-"
            c = ws.cell(row=row_cursor, column=3, value=str)
            c.alignment = Alignment(horizontal='left',vertical='center')                      
            try:
                str = hit["_source"]["message"].strip()
            except:
                str = u"-"
            c = ws.cell(row=row_cursor, column=4, value=str)
            c.alignment = Alignment(horizontal='left',vertical='center', wrap_text=True)
            
            urlstring = 'http://elastic/app/kibana#/doc/winlogbeat-*/' + hit["_index"] + u'/winlogbeat?id=' + hit["_id"]  
            c = ws.cell(row=row_cursor, column=5, value=u'Подробнее ...')            
            c.style = 'Hyperlink'
            c.alignment = Alignment(horizontal='left',vertical='center')
            c.hyperlink = (urlstring)
            
            row_cursor += 1
    except Exception as e: 
        print "ERROR: " +  unicode(e)
        pass
    c = ws.cell(row=row_cursor, column=1, value=u'Проверено: Событий, требующих внимания, нет.')
    c.alignment = Alignment(horizontal='left',vertical='center', wrap_text=True)
    c.font = Font(bold=True)
    ws.merge_cells(start_row=row_cursor,start_column=1,end_row=row_cursor+4,end_column=5)

if ips :
    if first :
        ws = wb.active
        ws.title = u'IPS'
        first = False
    else : 
        ws = wb.create_sheet(u'IPS') 
    row_cursor = 1
    tabheader = u"Подозрительные события Fortigate с " + startdate + u" по " + enddate
    c = ws.cell(row=row_cursor, column=1, value=tabheader)
    c.style = 'Title'
    c.alignment = Alignment(horizontal='left',vertical='center')
    ws.merge_cells(start_row=row_cursor,start_column=1,end_row=row_cursor+2,end_column=6)
    row_cursor += 3
    colheader = [u'Сигнатура',u'Время',u'Источник',u'Целевой хост',u'Подробности',u'Справка']
    i = 0
    for head in colheader:
        i += 1
        c = ws.cell(row=row_cursor, column=i, value=head)
        c.style = 'Headline 2'
    row_cursor += 1
    ws.column_dimensions['A'].width = 80
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 20
    myquery =   {"query":\
                        {\
                            "constant_score":{ "filter":{"term":{"subtype":"ips"}} }\
                        },\
                    "sort":{"@timestamp":{"order":"asc"}},\
                    "size":"10000"\
                }

    try:
        res = es.search(index=getindexes(es,"fortigate-",dates),body=myquery)
        for hit in res['hits']['hits']:
            try:
                str = hit["_source"]["attack"]
            except:
                str = u"-"                
            c = ws.cell(row=row_cursor, column=1, value=str)
            c.alignment = Alignment(horizontal='left',vertical='center')
            timestr = hit["_source"]["@timestamp"].replace("T"," ").split(".")[0]
            timeUTC = datetime.strptime(timestr, "%Y-%m-%d %H:%M:%S")
            timeUTC = timeUTC.replace(tzinfo=timezone('UTC'))
            timeLocal = timeUTC.astimezone(timezone('Europe/Moscow'))
            str = timeLocal.strftime("%Y-%m-%d %H:%M:%S")
            c = ws.cell(row=row_cursor, column=2, value=str)
            c.alignment = Alignment(horizontal='left',vertical='center')
            try:
                str = hit["_source"]["src_ip"]
            except:
                str = u"-"
            c = ws.cell(row=row_cursor, column=3, value=str)
            c.alignment = Alignment(horizontal='left',vertical='center')
            try:
                str = hit["_source"]["dst_ip"]
            except:
                str = u"-"
            c = ws.cell(row=row_cursor, column=4, value=str)
            c.alignment = Alignment(horizontal='left',vertical='center')
            urlstring = 'http://elastic/app/kibana#/doc/fortigate-*/' + hit["_index"] + u'/fortigate?id=' + hit["_id"]  
            c = ws.cell(row=row_cursor, column=5, value=u'Подробнее ...')            
            c.style = 'Hyperlink'
            c.alignment = Alignment(horizontal='left',vertical='center')
            c.hyperlink = (urlstring)  
            try:
                str = hit["_source"]["ref"]
            except:
                str = u"-"
            if str != "-" :
                c = ws.cell(row=row_cursor, column=6, value=u'Справка...')            
                c.style = 'Hyperlink'
                c.alignment = Alignment(horizontal='left',vertical='center')
                c.hyperlink = (str)
            row_cursor += 1
    except Exception as e: 
        print "ERROR: " +  unicode(e)
        pass
            
    row_cursor += 2

    tabheader = u"Подозрительные события Suricata с " + startdate + u" по " + enddate
    c = ws.cell(row=row_cursor, column=1, value=tabheader)
    c.style = 'Title'
    c.alignment = Alignment(horizontal='left',vertical='center')
    ws.merge_cells(start_row=row_cursor,start_column=1,end_row=row_cursor+2,end_column=6)
    row_cursor += 3
    colheader = [u'Сигнатура',u'Время',u'Источник',u'Целевой хост',u'Подробности',u'Справка']
    i = 0
    for head in colheader:
        i += 1
        c = ws.cell(row=row_cursor, column=i, value=head)
        c.style = 'Headline 2'
    row_cursor += 1
    myquery =   {"query":\
        {\
            "constant_score":{\
                "filter":[\
                    {"bool":{\
                        "must":[\
                            {"bool": {\
                                "should":[\
                                    {"term":{"alert.category":"A Network Trojan was Detected"}},\
                                    {"term":{"alert.category":"Potentially Bad Traffic"}},\
                                    {"term":{"alert.category":"Potential Corporate Privacy Violation"}},\
                                    {"term":{"alert.category":"A Suspicious Filename was Detected"}},\
                                    {"term":{"alert.category":"Misc Attack"}},\
                                    {"term":{"alert.category":"Attempted Denial of Service"}},\
                                ],\
                                "must_not":[\
                                    {"term":{"alert.signature_id":"2014170"}},\
                                    {"term":{"alert.signature_id":"2002157"}},\
                                    {"term":{"alert.signature_id":"2001595"}},\
                                    {"term":{"src_ip":"172.17.2.14"}},\
                                    {"term":{"src_ip":"172.16.0.37"}},\
                                    {"term":{"src_ip":"172.16.0.38"}},\
                                ]\
                            }}\
                        ]\
                    }},\
                ]\
            }\
        },\
        "sort":[{"alert.signature":{"order":"asc"}}, {"@timestamp":{"order":"asc"}}],\
        "size":"10000"\
    }


    try:
        res = es.search(index=getindexes(es,"suricata-",dates),body=myquery)
        prev_str=""
        for hit in res['hits']['hits']:
            try:
                str = hit["_source"]["alert"]["signature"]
            except:
                str = u"-"
            if prev_str != "" :
                if prev_str != str :
                    prev_str = str
                    row_cursor += 1
            else :
                prev_str = str
            c = ws.cell(row=row_cursor, column=1, value=str)
            c.alignment = Alignment(horizontal='left',vertical='center')
            timestr = hit["_source"]["@timestamp"].replace("T"," ").split(".")[0]
            timeUTC = datetime.strptime(timestr, "%Y-%m-%d %H:%M:%S")
            timeUTC = timeUTC.replace(tzinfo=timezone('UTC'))
            timeLocal = timeUTC.astimezone(timezone('Europe/Moscow'))
            str = timeLocal.strftime("%Y-%m-%d %H:%M:%S")
            c = ws.cell(row=row_cursor, column=2, value=str)
            c.alignment = Alignment(horizontal='left',vertical='center')
            try:
                str = hit["_source"]["src_ip"]
            except:
                str = u"-"
            c = ws.cell(row=row_cursor, column=3, value=str)
            c.alignment = Alignment(horizontal='left',vertical='center')
            try:
                str = hit["_source"]["dest_ip"]
            except:
                str = u"-"
            c = ws.cell(row=row_cursor, column=4, value=str)
            c.alignment = Alignment(horizontal='left',vertical='center')
            urlstring = 'http://elastic/app/kibana#/doc/suricata-*/' + hit["_index"] + u'/suricata?id=' + hit["_id"]  
            c = ws.cell(row=row_cursor, column=5, value=u'Подробнее ...')            
            c.style = 'Hyperlink'
            c.alignment = Alignment(horizontal='left',vertical='center')
            c.hyperlink = (urlstring)
            row_cursor += 1
    except Exception as e: 
        print "ERROR: " +  unicode(e)
        pass
    c = ws.cell(row=row_cursor, column=1, value=u'Проверено: Событий, требующих внимания, нет.')
    c.alignment = Alignment(horizontal='left',vertical='center', wrap_text=True)
    c.font = Font(bold=True)
    ws.merge_cells(start_row=row_cursor,start_column=1,end_row=row_cursor+4,end_column=6)    

wb.save('/usr/local/reports/' + 'daily-' + today.strftime('%Y-%m-%d') + '.xlsx')
